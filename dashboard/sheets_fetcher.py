"""
Google Sheets 매출 데이터 수집 모듈

두 개의 시트에서 데이터를 가져옵니다:
1. [뉴믹스] 일간 보고 지표 템플릿 - 채널별 일간 매출/목표
2. (뉴마센) 데일리 국가별 매출 check - 국가별 방문객/매출 (성수+북촌)
"""

import sys
import csv
import io
import urllib.request
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent / "00-system" / "02-scripts" / "gws"))
from gws_auth import authenticate

DATA_DIR = Path(__file__).parent / "data"

# 시트 ID
DAILY_REPORT_ID = "1r4MOXIXJEL9qdO5FRaJfhmQV7xxqsB6_1Q5y33cSh-8"  # 일간 보고 지표
COUNTRY_SALES_ID = "1mPrRbn9iz93BTRRp0_ibbplh6NPMNpYYjebFWCCcAPI"  # 뉴마센 데일리 국가별


def _export_sheet_csv(sheet_id: str, gid: int = 0, use_drive_api: bool = False) -> str:
    """Google Sheet를 CSV로 export합니다.

    Args:
        sheet_id: 스프레드시트 ID
        gid: 시트 탭 ID (0 = 첫 번째 탭)
        use_drive_api: True면 Drive API export_media 사용 (첫 번째 활성 탭)
    """
    if use_drive_api:
        from googleapiclient.http import MediaIoBaseDownload
        service = _get_drive_service()
        request = service.files().export_media(fileId=sheet_id, mimeType="text/csv")
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return fh.getvalue().decode("utf-8")
    else:
        creds = authenticate()
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {creds.token}"})
        resp = urllib.request.urlopen(req)
        return resp.read().decode("utf-8")


def _get_drive_service():
    from gws_auth import get_service
    return get_service("drive")


def _parse_number(s: str) -> float:
    """쉼표, 따옴표 제거 후 숫자 변환."""
    if not s or s.strip() == "":
        return 0.0
    s = s.strip().replace('"', '').replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ============================================================
# 뉴마센 데일리 국가별 매출 파싱
# ============================================================

def _download_xlsx(sheet_id: str) -> bytes:
    """Google Sheet를 XLSX로 다운로드합니다."""
    from googleapiclient.http import MediaIoBaseDownload
    service = _get_drive_service()
    request = service.files().export_media(
        fileId=sheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return fh.getvalue()


def fetch_country_sales(months: int = 2) -> list[dict]:
    """뉴마센 데일리 국가별 매출 데이터를 가져옵니다 (모든 월별 탭).

    Args:
        months: 최근 몇 개월 탭을 가져올지 (기본 2 = 당월 + 전월)

    Returns:
        [{date, store, country, visitors, sales, marketing_note}, ...]
    """
    import openpyxl

    xlsx_data = _download_xlsx(COUNTRY_SALES_ID)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)

    # 월별 탭만 필터 (2025., 2026. 으로 시작하는 탭)
    month_tabs = [name for name in wb.sheetnames if name.startswith("202")]
    # 최근 순 정렬
    month_tabs.sort(reverse=True)
    target_tabs = month_tabs[:months]

    records = []

    for tab_name in target_tabs:
        ws = wb[tab_name]
        rows = list(ws.iter_rows(values_only=True))

        # 탭 이름에서 연도/월 추출 (예: "2026.03" → 2026, 3)
        import re
        tab_match = re.match(r"(\d{4})\.(\d{2})", tab_name)
        tab_year = int(tab_match.group(1)) if tab_match else None
        tab_month = int(tab_match.group(2)) if tab_match else None

        header_row = rows[0] if rows else []
        col_row = rows[1] if len(rows) > 1 else []

        bukchon_start = None
        for i, h in enumerate(header_row):
            if h and "북촌" in str(h):
                bukchon_start = i
                break

        other_note_idx = None
        for i, h in enumerate(col_row):
            if h and "기타" in str(h) and "특이" in str(h):
                other_note_idx = i
                break

        for row in rows[2:]:
            if not row or not row[0]:
                continue

            date_val = row[0]
            if hasattr(date_val, 'strftime'):
                # openpyxl이 읽은 datetime - 탭 이름 기준으로 연도 보정
                if tab_year and tab_month:
                    date_formatted = datetime(
                        tab_year, date_val.month, date_val.day
                    ).strftime("%Y-%m-%d")
                else:
                    date_formatted = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str) and "/" in date_val:
                try:
                    parts = date_val.split("/")
                    month, day = int(parts[0]), int(parts[1])
                    year = tab_year or (2026 if month <= 6 else 2025)
                    date_formatted = datetime(year, month, day).strftime("%Y-%m-%d")
                except (ValueError, IndexError):
                    continue
            else:
                continue

            # 수식 행 스킵 (=E3/D3 같은 것)
            if any(str(c or "").startswith("=") for c in row[3:9] if c):
                continue

            day_of_week = str(row[1] or "").strip().split("\n")[0]
            marketing_note = str(row[9] or "").strip() if len(row) > 9 else ""
            other_note = str(row[other_note_idx] or "").strip() if other_note_idx and len(row) > other_note_idx else ""

            # 성수점 데이터 (컬럼 3~8)
            seongsu_cols = [(3, 4, "일본"), (5, 6, "대만"), (7, 8, "중국")]
            for vis_idx, sales_idx, country in seongsu_cols:
                visitors = int(float(row[vis_idx] or 0)) if len(row) > vis_idx and row[vis_idx] else 0
                sales = int(float(row[sales_idx] or 0)) if len(row) > sales_idx and row[sales_idx] else 0
                if visitors > 0 or sales > 0:
                    records.append({
                        "date": date_formatted,
                        "day_of_week": day_of_week,
                        "store": "성수",
                        "country": country,
                        "visitors": visitors,
                        "sales": sales,
                        "marketing_note": marketing_note,
                        "other_note": other_note,
                    })

            # 북촌점 데이터
            if bukchon_start:
                bukchon_cols = [
                    (bukchon_start, bukchon_start + 1, "일본"),
                    (bukchon_start + 2, bukchon_start + 3, "대만"),
                    (bukchon_start + 4, bukchon_start + 5, "중국"),
                ]
                for vis_idx, sales_idx, country in bukchon_cols:
                    visitors = int(float(row[vis_idx] or 0)) if len(row) > vis_idx and row[vis_idx] else 0
                    sales = int(float(row[sales_idx] or 0)) if len(row) > sales_idx and row[sales_idx] else 0
                    if visitors > 0 or sales > 0:
                        records.append({
                            "date": date_formatted,
                            "day_of_week": day_of_week,
                            "store": "북촌",
                            "country": country,
                            "visitors": visitors,
                            "sales": sales,
                            "marketing_note": "",
                            "other_note": "",
                        })

    wb.close()
    records.sort(key=lambda x: x["date"])
    return records


# ============================================================
# 일간 보고 지표 파싱
# ============================================================

def fetch_daily_report() -> dict:
    """일간 보고 지표 데이터를 가져옵니다.

    Returns:
        {
            "targets": [{date, channel, target}, ...],
            "actuals": [{date, channel, sales}, ...],
            "summary": {total_target, total_actual, achievement_rate},
        }
    """
    raw = _export_sheet_csv(DAILY_REPORT_ID, use_drive_api=True)
    reader = csv.reader(io.StringIO(raw))
    rows = list(reader)

    # 날짜 행 (row 6, 0-indexed 5): 3/1, 3/2, ...
    date_row = rows[5] if len(rows) > 5 else []
    dates = []
    for i in range(3, len(date_row)):
        d = date_row[i].strip()
        if d and "/" in d:
            parts = d.split("/")
            try:
                month, day = int(parts[0]), int(parts[1])
                dates.append(datetime(2026, month, day).strftime("%Y-%m-%d"))
            except (ValueError, IndexError):
                dates.append("")
        else:
            dates.append("")

    # 매출 목표 (rows 7-12, 0-indexed 6-11)
    target_channels = {
        6: "성수점", 7: "북촌점", 8: "네이버",
        9: "쿠팡", 10: "컬리", 11: "기타",
    }
    targets = []
    for row_idx, channel in target_channels.items():
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        for i, date in enumerate(dates):
            if not date:
                continue
            val = _parse_number(row[i + 3]) if i + 3 < len(row) else 0
            if val > 0:
                targets.append({
                    "date": date,
                    "channel": channel,
                    "target": val,
                })

    # 실매출 (rows 16-21, 0-indexed 15-20)
    actual_channels = {
        15: "성수점", 16: "북촌점", 17: "네이버",
        18: "쿠팡", 19: "컬리", 20: "기타",
    }
    actuals = []
    for row_idx, channel in actual_channels.items():
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        for i, date in enumerate(dates):
            if not date:
                continue
            val = _parse_number(row[i + 3]) if i + 3 < len(row) else 0
            if val > 0:
                actuals.append({
                    "date": date,
                    "channel": channel,
                    "sales": val,
                })

    # 달성률
    total_target = _parse_number(rows[12][1]) if len(rows) > 12 else 0
    total_actual = _parse_number(rows[21][1]) if len(rows) > 21 else 0
    rate = (total_actual / total_target * 100) if total_target > 0 else 0

    return {
        "targets": targets,
        "actuals": actuals,
        "summary": {
            "total_target": total_target,
            "total_actual": total_actual,
            "achievement_rate": round(rate, 1),
        },
    }


# ============================================================
# 통합 데이터 생성
# ============================================================

def fetch_all_sales_data() -> dict:
    """두 시트에서 모든 매출 데이터를 가져옵니다.

    Returns:
        {
            "country_sales": [...],  # 국가별 매출 (뉴마센 데일리)
            "channel_report": {...}, # 채널별 보고 (일간 보고 지표)
        }
    """
    result = {"country_sales": [], "channel_report": {}}

    try:
        result["country_sales"] = fetch_country_sales()
        print(f"국가별 매출: {len(result['country_sales'])}건")
    except Exception as e:
        print(f"국가별 매출 수집 실패: {e}")

    try:
        result["channel_report"] = fetch_daily_report()
        print(f"채널별 보고: 목표 {len(result['channel_report'].get('targets', []))}건, "
              f"실적 {len(result['channel_report'].get('actuals', []))}건")
    except Exception as e:
        print(f"채널별 보고 수집 실패: {e}")

    return result


def save_country_sales_csv(records: list[dict]):
    """국가별 매출을 CSV로 저장합니다."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    filepath = DATA_DIR / "country_sales.csv"
    fields = ["date", "day_of_week", "store", "country", "visitors", "sales",
              "marketing_note", "other_note"]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)
    print(f"저장: {filepath} ({len(records)}건)")


def country_sales_to_dashboard_format(records: list[dict]) -> list[dict]:
    """국가별 매출 데이터를 대시보드 형식으로 변환합니다.

    성수+북촌 합산하여 국가별 일간 매출로 변환.
    """
    from collections import defaultdict
    daily = defaultdict(lambda: {"sales": 0, "visitors": 0})

    for r in records:
        key = (r["date"], r["country"])
        daily[key]["sales"] += r["sales"]
        daily[key]["visitors"] += r["visitors"]

    result = []
    for (date, country), vals in sorted(daily.items()):
        result.append({
            "date": date,
            "country": country,
            "sales": vals["sales"],
            "visitors": vals["visitors"],
        })
    return result


if __name__ == "__main__":
    print("=== 데이터 수집 시작 ===\n")
    data = fetch_all_sales_data()

    if data["country_sales"]:
        save_country_sales_csv(data["country_sales"])
        dashboard_data = country_sales_to_dashboard_format(data["country_sales"])
        print(f"\n대시보드용 데이터: {len(dashboard_data)}건")
        for d in dashboard_data[:5]:
            print(f"  {d['date']} | {d['country']} | 매출 {d['sales']:,}원 | 방문 {d['visitors']}명")

    if data["channel_report"]:
        s = data["channel_report"]["summary"]
        print(f"\n채널 보고 요약: 목표 {s['total_target']}만 / 실적 {s['total_actual']}만 / 달성률 {s['achievement_rate']}%")
